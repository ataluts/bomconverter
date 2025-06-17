import os
import openpyxl
import dict_locale as lcl
from typedef_cl import CL_typeDef                               #класс списка компонентов

log_indent = " " * 4

#Импортирует список компонентов из формата xlsx
def importz(address, **kwargs):
    print('INFO >> CL importing module running with parameters:')
    print(f"{log_indent * 3} input: {os.path.basename(address)}")

    #locale
    locale_index = kwargs.get('locale_index', lcl.LocaleIndex.RU.value)

    title_list_components = kwargs.get('title_list_components', lcl.build_cl.TITLE_COMPONENTS_LIST.value[locale_index])
    title_list_accessories = kwargs.get('title_list_accessories', lcl.build_cl.TITLE_ACCESSORIES_LIST.value[locale_index])
    title_list_substitutes = kwargs.get('title_list_substitutes', lcl.build_cl.TITLE_SUBSTITUTES_LIST.value[locale_index])
    format_groupvalue_delimiter = kwargs.get('format_groupvalue_delimiter', ', ')
    format_singlevalue_delimiter = kwargs.get('format_singlevalue_delimiter', '|')

    #читаем данные из файла
    if os.path.isfile(address):
        print("INFO >> Analyzing data in Excel file:")
        workbook = openpyxl.load_workbook(address)
        cl = CL_typeDef(workbook.properties.title)

        #определяем листы
        worksheet_components = None
        worksheet_accessories = None
        worksheet_substitutes = None
        for worksheet in workbook:
            if worksheet.title == title_list_components:
                worksheet_components = worksheet
                continue
            if worksheet.title == title_list_accessories:
                worksheet_accessories = worksheet
                continue
            if worksheet.title == title_list_substitutes:
                worksheet_substitutes = worksheet
                continue

        message = f"{log_indent * 3} worksheets found: "
        if worksheet_components is not None: message += f"Components='{title_list_components}', "
        if worksheet_accessories is not None: message += f"Accessories='{title_list_accessories}', "
        if worksheet_substitutes is not None: message += f"Substitutes='{title_list_substitutes}', "
        print(message.rstrip(', '))

        #лист с компонентами
        if worksheet_components is not None:
            print("INFO >> Processing components worksheet:")
            cl.components = CL_typeDef.Sublist(title_list_components)

            #определяем индексы по заголовку таблицы
            column_index_designator   = None    #Поз. обозначение
            column_index_kind         = None    #Тип элемента
            column_index_value        = None    #Номинал
            column_index_description  = None    #Описание
            column_index_package      = None    #Корпус
            column_index_manufacturer = None    #Производитель
            column_index_quantity     = None    #Кол-во
            column_index_note         = None    #Примечание
            for i, cell in enumerate(worksheet_components[1]):
                if   cell.value == lcl.export_cl_xlsx.HEADER_DESIGNATOR.value[locale_index]:     column_index_designator = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_COMPONENT_TYPE.value[locale_index]: column_index_kind = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_VALUE.value[locale_index]:          column_index_value = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_DESCRIPTION.value[locale_index]:    column_index_description = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_PACKAGE.value[locale_index]:        column_index_package = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_MANUFACTURER.value[locale_index]:   column_index_manufacturer = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_QUANTITY.value[locale_index]:       column_index_quantity = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_NOTE.value[locale_index]:           column_index_note = i

            message = f"{log_indent * 3} columns indexes: Designator={column_index_designator}, Kind={column_index_kind}, Value={column_index_value}, Description={column_index_description}, Package={column_index_package}, Manufacturer={column_index_manufacturer}, Quantity={column_index_quantity}, Note={column_index_note}"
            print(message)

            #читаем данные из таблицы перебирая строки
            print(f"{log_indent * 3} reading rows", end="... ")
            for row in worksheet_components.iter_rows(2, worksheet_components.max_row):
                entry_designator   = None
                entry_kind         = None
                entry_value        = None
                entry_description  = None
                entry_package      = None
                entry_manufacturer = None
                entry_quantity     = None
                entry_note         = None
                entry_flag         = None

                #заполняем данные
                if column_index_designator is not None   and row[column_index_designator].value   is not None: entry_designator   = row[column_index_designator].value.split(format_groupvalue_delimiter)
                if column_index_kind is not None         and row[column_index_kind].value         is not None: entry_kind         = row[column_index_kind].value.split(format_singlevalue_delimiter)
                if column_index_value is not None        and row[column_index_value].value        is not None: entry_value        = row[column_index_value].value.split(format_singlevalue_delimiter)
                if column_index_description is not None  and row[column_index_description].value  is not None: entry_description  = row[column_index_description].value.split(format_singlevalue_delimiter)
                if column_index_package is not None      and row[column_index_package].value      is not None: entry_package      = row[column_index_package].value.split(format_singlevalue_delimiter)
                if column_index_manufacturer is not None and row[column_index_manufacturer].value is not None: entry_manufacturer = row[column_index_manufacturer].value.split(format_singlevalue_delimiter)
                if column_index_note is not None         and row[column_index_note].value         is not None: entry_note         = row[column_index_note].value.split(format_singlevalue_delimiter)
                entry_quantity, entry_flag = cellvalue_toint(row[column_index_quantity].value)
                
                entry = CL_typeDef.ComponentEntry(entry_designator, entry_kind, entry_value, entry_description, entry_package, entry_manufacturer, entry_quantity, entry_note, entry_flag)
                entry.check()

                if worksheet_accessories is not None:
                    #есть отдельный лист с аксессуарами -> добавляем все записи в список компонентов
                    cl.components.entries.append(entry)
                else:
                    #нет отдельного листа с аксессуарами -> добавляем записи с десигнаторами в компоненты, а без десигнаторов в аксессуары
                    if len(entry.designator) > 0:
                        cl.components.entries.append(entry)
                    else:
                        if cl.accessories is None: cl.accessories = CL_typeDef.Sublist(title_list_accessories)
                        cl.accessories.entries.append(entry)
            
            entry_num = len(cl.components.entries)
            if cl.accessories is not None: entry_num += len(cl.accessories.entries)
            print(f"done. ({entry_num} entries)")

        #лист с аксессуарами
        if worksheet_accessories is not None:
            print("INFO >> Processing accessories worksheet:")
            cl.accessories = CL_typeDef.Sublist(title_list_accessories)

            #определяем индексы по заголовку таблицы
            column_index_designator   = None    #Поз. обозначение
            column_index_kind         = None    #Тип элемента
            column_index_value        = None    #Номинал
            column_index_description  = None    #Описание
            column_index_package      = None    #Корпус
            column_index_manufacturer = None    #Производитель
            column_index_quantity     = None    #Кол-во
            column_index_note         = None    #Примечание
            for i, cell in enumerate(worksheet_accessories[1]):
                if   cell.value == lcl.export_cl_xlsx.HEADER_DESIGNATOR.value[locale_index]:     column_index_designator = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_COMPONENT_TYPE.value[locale_index]: column_index_kind = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_VALUE.value[locale_index]:          column_index_value = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_DESCRIPTION.value[locale_index]:    column_index_description = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_PACKAGE.value[locale_index]:        column_index_package = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_MANUFACTURER.value[locale_index]:   column_index_manufacturer = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_QUANTITY.value[locale_index]:       column_index_quantity = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_NOTE.value[locale_index]:           column_index_note = i

            message = f"{log_indent * 3} columns indexes: Designator={column_index_designator}, Kind={column_index_kind}, Value={column_index_value}, Description={column_index_description}, Package={column_index_package}, Manufacturer={column_index_manufacturer}, Quantity={column_index_quantity}, Note={column_index_note}"
            print(message)

            #читаем данные из таблицы перебирая строки
            print(f"{log_indent * 3} reading rows", end="... ")
            for row in worksheet_accessories.iter_rows(2, worksheet_accessories.max_row):
                entry_designator   = None
                entry_kind         = None
                entry_value        = None
                entry_description  = None
                entry_package      = None
                entry_manufacturer = None
                entry_quantity     = None
                entry_note         = None
                entry_flag         = None

                #заполняем данные
                if column_index_designator is not None   and row[column_index_designator].value   is not None: entry_designator   = row[column_index_designator].value.split(format_groupvalue_delimiter)
                if column_index_kind is not None         and row[column_index_kind].value         is not None: entry_kind         = row[column_index_kind].value.split(format_singlevalue_delimiter)
                if column_index_value is not None        and row[column_index_value].value        is not None: entry_value        = row[column_index_value].value.split(format_singlevalue_delimiter)
                if column_index_description is not None  and row[column_index_description].value  is not None: entry_description  = row[column_index_description].value.split(format_singlevalue_delimiter)
                if column_index_package is not None      and row[column_index_package].value      is not None: entry_package      = row[column_index_package].value.split(format_singlevalue_delimiter)
                if column_index_manufacturer is not None and row[column_index_manufacturer].value is not None: entry_manufacturer = row[column_index_manufacturer].value.split(format_singlevalue_delimiter)
                if column_index_note is not None         and row[column_index_note].value         is not None: entry_note         = row[column_index_note].value.split(format_singlevalue_delimiter)
                entry_quantity, entry_flag = cellvalue_toint(row[column_index_quantity].value)
               
                entry = CL_typeDef.ComponentEntry(entry_designator, entry_kind, entry_value, entry_description, entry_package, entry_manufacturer, entry_quantity, entry_note, entry_flag)
                entry.check()
                cl.accessories.entries.append(entry)

            print(f"done. ({len(cl.accessories.entries)} entries)")

        #лист со списком замен
        if worksheet_substitutes is not None:
            print("INFO >> Processing substitutes worksheet:")
            cl.substitutes = CL_typeDef.Sublist(title_list_substitutes)

            #определяем индексы по заголовку таблицы
            column_index_entry_value        = None    #Изнач. номинал
            column_index_entry_manufacturer = None    #Изнач. производитель
            column_index_entry_quantity     = None    #Изнач. кол-во
            column_index_group_designator   = None    #Поз. обозначение
            column_index_group_quantity     = None    #Зам. кол-во
            column_index_subst_value        = None    #Зам. номинал
            column_index_subst_manufacturer = None    #Зам. производитель
            column_index_subst_note         = None    #Зам. примечание
            for i, cell in enumerate(worksheet_substitutes[1], 1):
                if   cell.value == lcl.export_cl_xlsx.HEADER_ORIGINAL_VALUE.value[locale_index]:          column_index_entry_value = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_ORIGINAL_MANUFACTURER.value[locale_index]:   column_index_entry_manufacturer = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_ORIGINAL_QUANTITY.value[locale_index]:       column_index_entry_quantity = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_DESIGNATOR.value[locale_index]:              column_index_group_designator = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_SUBSTITUTE_QUANTITY.value[locale_index]:     column_index_group_quantity = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_SUBSTITUTE_VALUE.value[locale_index]:        column_index_subst_value = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_SUBSTITUTE_MANUFACTURER.value[locale_index]: column_index_subst_manufacturer = i
                elif cell.value == lcl.export_cl_xlsx.HEADER_SUBSTITUTE_NOTE.value[locale_index]:         column_index_subst_note = i

            message = f"{log_indent * 3} columns indexes: Entry.Value={column_index_entry_value}, Entry.Manufacturer={column_index_entry_manufacturer}, Entry.Quantity={column_index_entry_quantity}, Group.Designator={column_index_group_designator}, Group.Quantity={column_index_group_quantity}, Substitute.Value={column_index_subst_value}, Substitute.Manufacturer={column_index_subst_manufacturer}, Substitute.Note={column_index_subst_note}"
            print(message)

            print(f"{log_indent * 3} reading rows", end="... ")
            counter_entries = 0
            counter_groups = 0
            counter_substitutes = 0
            row_index = 2
            while row_index <= worksheet_substitutes.max_row:
                #элемент замены
                entry_rows, entry_cols = cell_get_size(worksheet_substitutes, worksheet_substitutes.cell(row = row_index, column = column_index_entry_value))
                entry_maxrow = row_index + entry_rows - 1
                entry_value = cell_get_merged_value(worksheet_substitutes, worksheet_substitutes.cell(row = row_index, column = column_index_entry_value))
                entry_manufacturer = cell_get_merged_value(worksheet_substitutes, worksheet_substitutes.cell(row = row_index, column = column_index_entry_manufacturer))
                entry_quantity = cell_get_merged_value(worksheet_substitutes, worksheet_substitutes.cell(row = row_index, column = column_index_entry_quantity))
                entry_quantity, entry_flag = cellvalue_toint(entry_quantity)
                entry = CL_typeDef.SubstituteEntry(entry_value, entry_manufacturer, entry_quantity, [], entry_flag)
                cl.substitutes.entries.append(entry)
                counter_entries += 1
                #заменные группы
                while row_index <= entry_maxrow:
                    group_rows, group_cols = cell_get_size(worksheet_substitutes, worksheet_substitutes.cell(row = row_index, column = column_index_group_designator))
                    group_maxrow = row_index + group_rows - 1
                    group_designator = cell_get_merged_value(worksheet_substitutes, worksheet_substitutes.cell(row = row_index, column = column_index_group_designator))
                    if group_designator is not None: group_designator = group_designator.split(format_groupvalue_delimiter)
                    group_quantity = cell_get_merged_value(worksheet_substitutes, worksheet_substitutes.cell(row = row_index, column = column_index_group_quantity))
                    group_quantity, group_flag = cellvalue_toint(entry_quantity)
                    group = CL_typeDef.SubstituteEntry.SubstituteGroup(group_designator, group_quantity, [], group_flag)
                    entry.substitute_group.append(group)
                    counter_groups += 1
                    #замены в группе
                    while row_index <= group_maxrow:
                        substitute_value = worksheet_substitutes.cell(row = row_index, column = column_index_subst_value).value
                        substitute_manufacturer = worksheet_substitutes.cell(row = row_index, column = column_index_subst_manufacturer).value
                        substitute_note = worksheet_substitutes.cell(row = row_index, column = column_index_subst_note).value
                        substitute = CL_typeDef.SubstituteEntry.SubstituteGroup.Substitute(substitute_value, substitute_manufacturer, substitute_note)
                        group.substitute.append(substitute)
                        counter_substitutes += 1
                        row_index += 1
            print(f"done. ({counter_entries} entries, {counter_groups} groups, {counter_substitutes} substitutes)")

        components_quantity = 0
        if cl.components is not None: components_quantity = len(cl.components.entries)
        accessories_quantity = 0
        if cl.accessories is not None: accessories_quantity = len(cl.accessories.entries)
        substitutes_quantity = 0
        if cl.substitutes is not None: substitutes_quantity = len(cl.substitutes.entries)
        print(f"INFO >> CL imported. ({components_quantity} components, {accessories_quantity} accessories, {substitutes_quantity} substitutes)")
    else:
        print("ERROR >> file doesn't exist")

    print('INFO >> CL import finished.')
    return cl 

#Конвертирует значение ячейки в целочисленное и возвращает флаг
def cellvalue_toint(cell_value):
    try:
        value = int(cell_value)
        flag = None
    except (ValueError, TypeError):
        value = 0
        flag = CL_typeDef.FlagType.ERROR
    return value, flag

#Возвращает размер ячейки
def cell_get_size(worksheet, cell):
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            rows = merged_range.size["rows"]
            cols = merged_range.size["columns"]
            return rows, cols
    return 1, 1

#Возвращает значение объединённой ячейки
def cell_get_merged_value(worksheet, cell):
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left = worksheet[merged_range.min_row][merged_range.min_col - 1]
            return top_left.value
    return cell.value